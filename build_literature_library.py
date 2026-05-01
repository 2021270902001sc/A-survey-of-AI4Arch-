import csv
import os
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
from pathlib import Path

ROOT = Path("AI_for_Arch_LLM_Literature")

CATEGORIES = {
    "01_LLM_Agents_for_DSE": "LLM/agent-guided design space exploration",
    "02_LLM_for_Microarchitecture_Policy_Design": "Microarchitecture policy design and reasoning",
    "03_LLM_for_Simulation_Modeling_and_gem5": "Architecture simulation, modeling, gem5/ChampSim workflows",
    "04_LLM_for_Performance_Analysis_and_Optimization": "Performance analysis and optimization",
    "05_LLM_for_Hardware_Spec_RTL_Verification_when_arch_relevant": "HLS/RTL/specification/verification when architecture-relevant",
    "06_LLM_for_HPC_System_Codesign": "HPC/system/software-hardware co-design",
    "90_Surveys_Position_and_Vision": "Surveys, benchmarks, position papers, vision papers",
    "99_Candidates_Borderline": "Borderline EDA/RTL/compiler candidates",
}

FIELDS = [
    "id",
    "status",
    "title",
    "authors",
    "year",
    "date",
    "venue",
    "paper_type",
    "direction",
    "technical_route",
    "llm_role",
    "architecture_target",
    "simulator_toolchain",
    "evaluation_metrics",
    "category",
    "recent_half_year_arxiv",
    "pdf_path",
    "official_url",
    "oa_url",
    "doi_arxiv_id",
    "notes",
]

ENTRIES = [
    {
        "id": "pf_llm_2026",
        "status": "main",
        "title": "PF-LLM: Large Language Model Hinted Hardware Prefetching",
        "authors": "Ceyu Xu; Xiangfeng Sun; Weihang Li; Chen Bai; Bangyan Wang; Mengming Li; Zhiyao Xie; Yuan Xie",
        "year": "2026",
        "date": "2026-03-26",
        "venue": "ASPLOS 2026",
        "paper_type": "top conference paper",
        "direction": "LLM-assisted microarchitecture policy design",
        "technical_route": "LLM-assisted microarchitecture policy design",
        "llm_role": "LLM hints hardware prefetching policy/design choices",
        "architecture_target": "hardware prefetching / processor microarchitecture",
        "simulator_toolchain": "not available from public program page",
        "evaluation_metrics": "9.8% average IPC improvement over state-of-the-art hardware prefetching baselines on memory-intensive SPEC 2017; 18.9% over state-of-the-art ensemble methods",
        "category": "02_LLM_for_Microarchitecture_Policy_Design",
        "recent_half_year_arxiv": "unknown",
        "official_url": "https://doi.org/10.1145/3779212.3790202",
        "oa_url": "https://fact-lab.hkust.edu.hk/publications/conference-paper/2025/xu-2025-pf-llm/3779212.3790202.pdf",
        "doi_arxiv_id": "DOI:10.1145/3779212.3790202",
        "notes": "Accepted in ASPLOS 2026 Session 7D Processor Microarchitecture; ASPLOS'26 Best Paper Award. Strong top-venue microarchitecture match.",
    },
    {
        "id": "cachemind_2026",
        "status": "main",
        "title": "CacheMind: From Miss Rates to Why -- Natural-Language, Trace-Grounded Reasoning for Cache Replacement",
        "authors": "Kaushal Mhapsekar; Azam Ghanbari; Bita Aslrousta; Samira Mirbagher-Ajorpaz",
        "year": "2026",
        "date": "2026-02-12",
        "venue": "ASPLOS 2026 / arXiv",
        "paper_type": "top conference paper / preprint",
        "direction": "LLM for cache replacement reasoning",
        "technical_route": "RAG/trace-grounded reasoning",
        "llm_role": "Conversational RAG system explains cache traces and replacement behavior",
        "architecture_target": "cache replacement",
        "simulator_toolchain": "ChampSim traces; SIEVE/RANGER retrievers",
        "evaluation_metrics": "CacheMindBench accuracy: trace-grounded and policy-specific reasoning tasks",
        "category": "02_LLM_for_Microarchitecture_Policy_Design",
        "recent_half_year_arxiv": "yes",
        "official_url": "https://doi.org/10.1145/3779212.3790136",
        "oa_url": "https://arxiv.org/abs/2602.12422",
        "doi_arxiv_id": "arXiv:2602.12422; DOI:10.1145/3779212.3790136",
        "notes": "Strong match: LLM/RAG directly supports microarchitectural cache-replacement analysis.",
    },
    {
        "id": "agentic_architect_2026",
        "status": "main",
        "title": "Agentic Architect: An Agentic AI Framework for Architecture Design Exploration and Optimization",
        "authors": "Alexander Blasberg; Vasilis Kypriotis; Dimitrios Skarlatos",
        "year": "2026",
        "date": "2026-04-28",
        "venue": "arXiv",
        "paper_type": "recent preprint",
        "direction": "agentic microarchitecture design",
        "technical_route": "Agentic closed-loop design",
        "llm_role": "LLM agent evolves code/design policies under simulator feedback",
        "architecture_target": "cache replacement; data prefetching; branch prediction",
        "simulator_toolchain": "cycle-accurate simulation; open-source agentic evaluation harness",
        "evaluation_metrics": "geomean IPC speedup over LRU/Mockingjay, Bimodal/Hashed Perceptron, no-prefetch/SMS",
        "category": "02_LLM_for_Microarchitecture_Policy_Design",
        "recent_half_year_arxiv": "yes",
        "official_url": "https://arxiv.org/abs/2604.25083",
        "oa_url": "https://arxiv.org/abs/2604.25083",
        "doi_arxiv_id": "arXiv:2604.25083",
        "notes": "Recent half-year arXiv; one of the clearest end-to-end LLM-agent microarchitecture design papers.",
    },
    {
        "id": "gem5_copilot_2025",
        "status": "main",
        "title": "gem5 Co-Pilot: AI Assistant Agent for Architectural Design Space Exploration",
        "authors": "Zuoming Fu; Alexander M. Manley; Mohammad Alian",
        "year": "2025",
        "date": "2025-10-21",
        "venue": "CAMS 2025 / arXiv",
        "paper_type": "workshop paper / preprint",
        "direction": "LLM agent for gem5 DSE",
        "technical_route": "Natural-language DSE interface; Agentic closed-loop design; RAG/DSDB",
        "llm_role": "LLM state-machine agent analyzes past simulations, proposes gem5 parameter batches, and answers user questions",
        "architecture_target": "gem5 L2 cache hierarchy parameter DSE under cost/power constraints",
        "simulator_toolchain": "gem5; McPAT; Design Space Declarative Language; Design Space Database; Streamlit UI",
        "evaluation_metrics": "near-optimal perf_ratio with 2-12 gem5 runs; API cost below 0.5 USD/session",
        "category": "01_LLM_Agents_for_DSE",
        "recent_half_year_arxiv": "no (submitted 2025-10-21, nine days before cutoff)",
        "official_url": "https://arxiv.org/abs/2510.19577",
        "oa_url": "https://wisl.ece.cornell.edu/papers/gem5-co-pilot-cams-25.pdf",
        "doi_arxiv_id": "arXiv:2510.19577",
        "notes": "Directly matches the user request; important gem5-centered exemplar.",
    },
    {
        "id": "quarch_2025",
        "status": "main",
        "title": "QuArch: A Benchmark for Evaluating LLM Reasoning in Computer Architecture",
        "authors": "Shvetank Prakash; Andrew Cheng; Arya Tschand; Mark Mazumder; Varun Gohil; Jeffrey Ma; Jason Yik; Zishen Wan; Jessica Quaye; Elisavet Lydia Alvanaki; Avinash Kumar; Chandrashis Mazumdar; Tuhin Khare; Alexander Ingare; Ikechukwu Uchendu; Radhika Ghosal; Abhishek Tyagi; Chenyu Wang; Andrea Mattia Garavagno; Shuning Gu; Alicia Guo; Grace Hur; Luca Marcello Carloni; Tushar Krishna; Ankita Nayak; Amir Yazdanbakhsh; Vijay Janapa Reddi",
        "year": "2025",
        "date": "2025-10-24",
        "venue": "OpenReview / arXiv, submitted to ICLR 2026",
        "paper_type": "benchmark / preprint",
        "direction": "LLM benchmark for computer architecture reasoning",
        "technical_route": "Benchmark/dataset for architecture reasoning",
        "llm_role": "Evaluates LLM architecture knowledge, analysis, design, and implementation reasoning",
        "architecture_target": "processor design; memory systems; interconnection networks",
        "simulator_toolchain": "QA benchmark; LLM-as-judge and human grading",
        "evaluation_metrics": "2,671 expert-validated QA pairs; frontier model accuracy range about 34%-72%",
        "category": "90_Surveys_Position_and_Vision",
        "recent_half_year_arxiv": "no (submitted 2025-10-24, six days before cutoff)",
        "official_url": "https://openreview.net/forum?id=nhcz0uni55",
        "oa_url": "https://arxiv.org/abs/2510.22087",
        "doi_arxiv_id": "arXiv:2510.22087",
        "notes": "Not a design method, but foundational for evaluating LLMs in computer architecture.",
    },
    {
        "id": "lpcm_2025",
        "status": "main",
        "title": "Large Processor Chip Model",
        "authors": "Kaiyan Chang; Mingzhi Chen; Yunji Chen; Zhirong Chen; Dongrui Fan; Junfeng Gong; Nan Guo; Yinhe Han; Qinfen Hao; Shuo Hou; Xuan Huang; Pengwei Jin; Changxin Ke; Cangyuan Li; Guangli Li; Huawei Li; Kuan Li; Naipeng Li; Shengwen Liang; Cheng Liu; Hongwei Liu; Jiahua Liu; Junliang Lv; Jianan Mu; Jin Qin; Bin Sun; Chenxi Wang; Duo Wang; Mingjun Wang; Ying Wang; Chenggang Wu; Peiyang Wu; Teng Wu; Xiao Xiao; Mengyao Xie; Chenwei Xiong; Ruiyuan Xu; Mingyu Yan; Xiaochun Ye; Kuai Yu; Rui Zhang; Shuoming Zhang; Jiacheng Zhao et al.",
        "year": "2025",
        "date": "2025-06-03",
        "venue": "arXiv / Science China Information Sciences for review",
        "paper_type": "position paper",
        "direction": "LLM-driven end-to-end processor/chip design vision",
        "technical_route": "Survey/vision; Agentic closed-loop design",
        "llm_role": "LPCM proposes LLM/agent levels for compiler, binary translation, simulator, partitioning, DSE, RTL generation",
        "architecture_target": "processor/chip architecture across software-hardware stack",
        "simulator_toolchain": "LLVM; gem5; Chisel; Verilog simulators; 3D Gaussian Splatting workload case study",
        "evaluation_metrics": "case-study effectiveness; mostly roadmap/position analysis",
        "category": "90_Surveys_Position_and_Vision",
        "recent_half_year_arxiv": "no",
        "official_url": "https://arxiv.org/abs/2506.02929",
        "oa_url": "https://arxiv.org/abs/2506.02929",
        "doi_arxiv_id": "arXiv:2506.02929",
        "notes": "Useful Chinese-led vision paper for LLM-for-architecture framing.",
    },
    {
        "id": "llm_dse_2025",
        "status": "main",
        "title": "LLM-DSE: Searching Accelerator Parameters with LLM Agents",
        "authors": "Hanyu Wang; Xinrui Wu; Zijian Ding; Su Zheng; Chengyue Wang; Tony Nowatzki; Yizhou Sun; Jason Cong",
        "year": "2025",
        "date": "2025-05-18",
        "venue": "arXiv",
        "paper_type": "preprint",
        "direction": "LLM agents for accelerator/HLS DSE",
        "technical_route": "Agentic closed-loop design",
        "llm_role": "Router, Specialist, Arbitrator, and Critic agents guide parameter search with online verbal learning",
        "architecture_target": "HLS directive/parameter optimization for domain-specific accelerators",
        "simulator_toolchain": "Merlin/Stratus/Vitis-style HLS flows; HLSyn; Rosetta",
        "evaluation_metrics": "2.55x performance gain over AutoDSE; runtime reduction; ablations on agent interactions",
        "category": "01_LLM_Agents_for_DSE",
        "recent_half_year_arxiv": "no",
        "official_url": "https://arxiv.org/abs/2505.12188",
        "oa_url": "https://arxiv.org/abs/2505.12188",
        "doi_arxiv_id": "arXiv:2505.12188",
        "notes": "Strong DSE fit; accelerator/HLS rather than CPU microarchitecture.",
    },
    {
        "id": "lift_2025",
        "status": "main",
        "title": "LIFT: LLM-Based Pragma Insertion for HLS via GNN Supervised Fine-Tuning",
        "authors": "Neha Prakriya; Zijian Ding; Yizhou Sun; Jason Cong",
        "year": "2025",
        "date": "2025-04-29",
        "venue": "arXiv",
        "paper_type": "preprint",
        "direction": "LLM-assisted HLS pragma generation",
        "technical_route": "Hardware design/spec optimization with LLM fine-tuning",
        "llm_role": "Fine-tuned LLM generates performance-critical pragmas with GNN supervision",
        "architecture_target": "FPGA/HLS microarchitecture transformations",
        "simulator_toolchain": "HLSyn; AutoDSE; HARP; ProGraML/LLVM IR graph supervision",
        "evaluation_metrics": "3.52x over AutoDSE; 2.16x over HARP; 66x over GPT-4o baseline",
        "category": "05_LLM_for_Hardware_Spec_RTL_Verification_when_arch_relevant",
        "recent_half_year_arxiv": "no",
        "official_url": "https://arxiv.org/abs/2504.21187",
        "oa_url": "https://arxiv.org/abs/2504.21187",
        "doi_arxiv_id": "arXiv:2504.21187",
        "notes": "HLS rather than classic Arch venue, but highly relevant to architecture-generation workflow.",
    },
    {
        "id": "idse_2025",
        "status": "main",
        "title": "iDSE: Navigating Design Space Exploration in High-Level Synthesis Using LLMs",
        "authors": "Runkai Li; Jia Xiong; Xi Wang",
        "year": "2025",
        "date": "2025-05-28",
        "venue": "arXiv",
        "paper_type": "preprint",
        "direction": "LLM-aided HLS DSE",
        "technical_route": "Natural-language/DSE reasoning; Agentic closed-loop design",
        "llm_role": "LLM prunes design space, generates warm-start samples, reflects on QoR trajectories, and refines Pareto designs",
        "architecture_target": "HLS directive configurations and generated microarchitectures",
        "simulator_toolchain": "vendor HLS synthesis flow; NSGA-II/ACO/MOEA-D/Lattice/HGBO-DSE baselines",
        "evaluation_metrics": "5.1x-16.6x better proximity to reference Pareto front; matches NSGA-II with 4.6% explored designs",
        "category": "01_LLM_Agents_for_DSE",
        "recent_half_year_arxiv": "no",
        "official_url": "https://arxiv.org/abs/2505.22086",
        "oa_url": "https://arxiv.org/abs/2505.22086",
        "doi_arxiv_id": "arXiv:2505.22086",
        "notes": "Important HLS DSE paper; likely arXiv-only in this pass.",
    },
    {
        "id": "mpm_llm4dse_2026",
        "status": "main",
        "title": "MPM-LLM4DSE: Reaching the Pareto Frontier in HLS with Multimodal Learning and LLM-Driven Exploration",
        "authors": "Lei Xu; Shanshan Wang; Chenglong Xiao",
        "year": "2026",
        "date": "2026-01-08",
        "venue": "arXiv",
        "paper_type": "recent preprint",
        "direction": "LLM-driven HLS Pareto DSE",
        "technical_route": "LLM optimizer with multimodal QoR predictor",
        "llm_role": "LLM acts as optimizer guided by pragma-impact prompt engineering; code text embeddings augment prediction",
        "architecture_target": "HLS pragma configuration / accelerator design",
        "simulator_toolchain": "HLS synthesis data; multimodal prediction model; ProgSG baseline",
        "evaluation_metrics": "up to 10.25x QoR-prediction improvement; 39.90% DSE performance gain over prior methods",
        "category": "01_LLM_Agents_for_DSE",
        "recent_half_year_arxiv": "yes",
        "official_url": "https://arxiv.org/abs/2601.04801",
        "oa_url": "https://arxiv.org/abs/2601.04801",
        "doi_arxiv_id": "arXiv:2601.04801",
        "notes": "Recent half-year arXiv; HLS-specific but strongly DSE-oriented.",
    },
    {
        "id": "diffhls_2026",
        "status": "main",
        "title": "DiffHLS: Differential Learning for High-Level Synthesis QoR Prediction with GNNs and LLM Code Embeddings",
        "authors": "Zedong Peng; Zeju Li; Qiang Xu; Jieru Zhao",
        "year": "2026",
        "date": "2026-04-10",
        "venue": "arXiv",
        "paper_type": "recent preprint",
        "direction": "LLM code embeddings for HLS QoR prediction",
        "technical_route": "Surrogate/QoR prediction augmented by pretrained code LLM embeddings",
        "llm_role": "Pretrained code LLM embeddings improve delta-path QoR prediction",
        "architecture_target": "HLS pragma-driven design variants",
        "simulator_toolchain": "PolyBench; ForgeHLS; GNN backbones",
        "evaluation_metrics": "lower MAPE than GNN baselines; LLM embeddings improve GNN-only ablation",
        "category": "05_LLM_for_Hardware_Spec_RTL_Verification_when_arch_relevant",
        "recent_half_year_arxiv": "yes",
        "official_url": "https://arxiv.org/abs/2604.09240",
        "oa_url": "https://arxiv.org/abs/2604.09240",
        "doi_arxiv_id": "arXiv:2604.09240",
        "notes": "Recent half-year arXiv; more prediction/surrogate than agentic design.",
    },
    {
        "id": "chathls_2025",
        "status": "main",
        "title": "ChatHLS: Towards Systematic Design Automation and Optimization for High-Level Synthesis",
        "authors": "Runkai Li; Jia Xiong; Xiuyuan He; Jiaqi Lv; Jieru Zhao; Qiang Xu; Xi Wang",
        "year": "2025",
        "date": "2025-07-01",
        "venue": "arXiv",
        "paper_type": "preprint",
        "direction": "multi-agent HLS design automation",
        "technical_route": "LLM for simulation workflow; hardware design/spec optimization",
        "llm_role": "Fine-tuned LLMs in multi-agent framework for HLS-C debugging and QoR-aware directive optimization",
        "architecture_target": "HLS code and accelerator design",
        "simulator_toolchain": "HLS toolchains; verification-oriented data augmentation",
        "evaluation_metrics": "82.7% repair pass rate; 1.9x-14.8x speedups; 4.9x geomean over DSL-based approaches",
        "category": "05_LLM_for_Hardware_Spec_RTL_Verification_when_arch_relevant",
        "recent_half_year_arxiv": "no",
        "official_url": "https://arxiv.org/abs/2507.00642",
        "oa_url": "https://arxiv.org/abs/2507.00642",
        "doi_arxiv_id": "arXiv:2507.00642",
        "notes": "Bridge between HLS code generation, repair, and architecture optimization.",
    },
    {
        "id": "hls_trans_2025",
        "status": "main",
        "title": "HLStrans: Dataset for C-to-HLS Hardware Code Synthesis",
        "authors": "Qingyun Zou; Nuo Chen; Yao Chen; Bingsheng He; Weng-Fai Wong",
        "year": "2025",
        "date": "2025-07-06",
        "venue": "OpenReview / arXiv, submitted to ICLR 2026",
        "paper_type": "dataset / benchmark",
        "direction": "dataset for LLM-driven C-to-HLS synthesis",
        "technical_route": "Benchmark/dataset; hardware design/spec optimization",
        "llm_role": "LLMs are benchmarked and used in augmentation pipeline with MCTS/DSE",
        "architecture_target": "FPGA/HLS program transformations and pragmas",
        "simulator_toolchain": "HLS synthesis annotations; testbenches; DSE pipeline",
        "evaluation_metrics": "124K+ paired C/HLS programs in initial report; retrieval/fine-tuning improves success and latency",
        "category": "05_LLM_for_Hardware_Spec_RTL_Verification_when_arch_relevant",
        "recent_half_year_arxiv": "no",
        "official_url": "https://openreview.net/forum?id=6UaXec8aUt",
        "oa_url": "https://arxiv.org/abs/2507.04315",
        "doi_arxiv_id": "arXiv:2507.04315",
        "notes": "Dataset/benchmark, included because it supports LLM-for-HLS architecture workflows.",
    },
    {
        "id": "hlspilot_2024",
        "status": "main",
        "title": "HLSPilot: LLM-based High-Level Synthesis",
        "authors": "Chenwei Xiong; Cheng Liu; Huawei Li; Xiaowei Li",
        "year": "2024",
        "date": "2024-08-13",
        "venue": "ICCAD 2024 / arXiv",
        "paper_type": "top EDA conference paper / preprint",
        "direction": "LLM-enabled end-to-end HLS hardware acceleration",
        "technical_route": "LLM for simulation workflow; hardware/software co-design; DSE tool use",
        "llm_role": "LLM profiles software, identifies bottlenecks, retrieves HLS strategies, generates optimized HLS code, invokes DSE tooling, and helps CPU-FPGA deployment",
        "architecture_target": "hybrid CPU-FPGA accelerators and HLS-generated microarchitectures",
        "simulator_toolchain": "gprof; Vitis HLS; Xilinx Alveo U280; GenHLSOptimizer; XRT",
        "evaluation_metrics": "comparable to or better than manual HLS designs on real-world application benchmarks",
        "category": "05_LLM_for_Hardware_Spec_RTL_Verification_when_arch_relevant",
        "recent_half_year_arxiv": "no",
        "official_url": "https://doi.org/10.1145/3676536.3676781",
        "oa_url": "https://arxiv.org/abs/2408.06810",
        "doi_arxiv_id": "arXiv:2408.06810; DOI:10.1145/3676536.3676781",
        "notes": "Early and important LLM-for-HLS framework; bridges profiling, partitioning, HLS codegen, DSE, and deployment.",
    },
    {
        "id": "hls_eval_2025",
        "status": "main",
        "title": "HLS-Eval: A Benchmark and Framework for Evaluating LLMs on High-Level Synthesis Design Tasks",
        "authors": "Stefan Abi-Karam; Cong Hao",
        "year": "2025",
        "date": "2025-04-16",
        "venue": "IEEE ICLAD 2025 / arXiv",
        "paper_type": "conference paper / benchmark",
        "direction": "benchmark for LLM-driven HLS design tasks",
        "technical_route": "Benchmark/dataset; hardware design/spec optimization",
        "llm_role": "Evaluates LLMs for HLS code generation and HLS-specific code edits targeting performance/hardware efficiency",
        "architecture_target": "HLS designs for accelerators and hardware systems",
        "simulator_toolchain": "Vitis HLS; automated parallel evaluation framework",
        "evaluation_metrics": "94 unique HLS designs; parseability, compilability, runnability, synthesizability, pass@k",
        "category": "05_LLM_for_Hardware_Spec_RTL_Verification_when_arch_relevant",
        "recent_half_year_arxiv": "no",
        "official_url": "https://doi.org/10.1109/ICLAD65226.2025.00021",
        "oa_url": "https://arxiv.org/abs/2504.12268",
        "doi_arxiv_id": "arXiv:2504.12268; DOI:10.1109/ICLAD65226.2025.00021",
        "notes": "Evaluation infrastructure rather than design optimizer, but central for measuring LLM-for-HLS progress.",
    },
    {
        "id": "sage_hls_2025",
        "status": "main",
        "title": "SAGE-HLS: Syntax-Aware AST-Guided LLM for High-Level Synthesis Code Generation",
        "authors": "M. Zafir Sadik Khan; Nowfel Mashnoor; Mohammad Akyash; Kimia Zamiri Azar; Hadi Kamali",
        "year": "2025",
        "date": "2025-08-05",
        "venue": "arXiv",
        "paper_type": "preprint",
        "direction": "syntax-aware LLM for HLS code generation",
        "technical_route": "Hardware design/spec optimization",
        "llm_role": "Fine-tuned LLM uses AST-guided instruction prompting and Verilog-to-C/C++ porting to improve synthesizable HLS code generation",
        "architecture_target": "HLS code for FPGA/ASIC hardware generation",
        "simulator_toolchain": "semi-automated HLS evaluation; AST-guided prompting",
        "evaluation_metrics": "synthesizability and functional correctness of generated HLS code",
        "category": "05_LLM_for_Hardware_Spec_RTL_Verification_when_arch_relevant",
        "recent_half_year_arxiv": "no",
        "official_url": "https://arxiv.org/abs/2508.03558",
        "oa_url": "https://arxiv.org/abs/2508.03558",
        "doi_arxiv_id": "arXiv:2508.03558",
        "notes": "HLS code-generation item; included because it targets synthesizable hardware design rather than generic code.",
    },
    {
        "id": "bench4hls_2026",
        "status": "main",
        "title": "Bench4HLS: End-to-End Evaluation of LLMs in High-Level Synthesis Code Generation",
        "authors": "M. Zafir Sadik Khan; Kimia Zamiri Azar; Hadi Kamali",
        "year": "2026",
        "date": "2026-01-16",
        "venue": "arXiv",
        "paper_type": "recent preprint / benchmark",
        "direction": "end-to-end benchmark for LLM-generated HLS",
        "technical_route": "Benchmark/dataset; hardware design/spec optimization",
        "llm_role": "Benchmarks LLM-generated HLS designs through compilation, simulation, synthesis, and PPA analysis",
        "architecture_target": "HLS designs from small kernels to complex accelerators",
        "simulator_toolchain": "Xilinx Vitis HLS; Catapult HLS validation; pluggable PPA API",
        "evaluation_metrics": "170 manually drafted and validated case studies; compilation, functional correctness, synthesis feasibility, PPA",
        "category": "05_LLM_for_Hardware_Spec_RTL_Verification_when_arch_relevant",
        "recent_half_year_arxiv": "yes",
        "official_url": "https://arxiv.org/abs/2601.19941",
        "oa_url": "https://arxiv.org/abs/2601.19941",
        "doi_arxiv_id": "arXiv:2601.19941",
        "notes": "Recent half-year arXiv benchmark; useful to compare HLS-Eval and HLStrans.",
    },
    {
        "id": "hls_directives_llm_2025",
        "status": "main",
        "title": "High-level Synthesis Directives Design Optimization via Large Language Model",
        "authors": "Xufeng Yao; Wenqian Zhao; Qi Sun; Cheng Zhuo; Bei Yu",
        "year": "2025",
        "date": "2025-07",
        "venue": "ACM Transactions on Design Automation of Electronic Systems",
        "paper_type": "journal article",
        "direction": "LLM for HLS directive optimization",
        "technical_route": "Hardware design/spec optimization; Agentic closed-loop design",
        "llm_role": "Uses LLMs as feature extractors and autonomous agents for HLS directive optimization",
        "architecture_target": "HLS directives and FPGA design space exploration",
        "simulator_toolchain": "HLS DSE workflow; Bayesian optimization context",
        "evaluation_metrics": "HLS directive optimization quality; see journal paper for details",
        "category": "05_LLM_for_Hardware_Spec_RTL_Verification_when_arch_relevant",
        "recent_half_year_arxiv": "no",
        "official_url": "https://doi.org/10.1145/3747291",
        "oa_url": "",
        "doi_arxiv_id": "DOI:10.1145/3747291",
        "notes": "Peer-reviewed ACM TODAES journal article; no open PDF found in this pass.",
    },
    {
        "id": "timelyhls_2025",
        "status": "main",
        "title": "TimelyHLS: LLM-Based Timing-Aware and Architecture-Specific FPGA HLS Optimization",
        "authors": "Nowfel Mashnoor; Mohammad Akyash; Hadi Kamali; Kimia Zamiri Azar",
        "year": "2025",
        "date": "2025-07-23",
        "venue": "IEEE COINS 2025 / arXiv",
        "paper_type": "conference paper / preprint",
        "direction": "RAG-assisted FPGA/HLS optimization",
        "technical_route": "RAG/trace-grounded reasoning; hardware design/spec optimization",
        "llm_role": "LLM with RAG generates and iteratively refines timing- and architecture-specific pragmas from synthesis feedback",
        "architecture_target": "FPGA-specific HLS code and RTL quality",
        "simulator_toolchain": "commercial HLS toolchains; structured FPGA architecture knowledge base",
        "evaluation_metrics": "up to 70% less manual tuning; up to 4x latency speedup; >50% area savings in cases",
        "category": "05_LLM_for_Hardware_Spec_RTL_Verification_when_arch_relevant",
        "recent_half_year_arxiv": "no",
        "official_url": "https://doi.org/10.48550/arXiv.2507.17962",
        "oa_url": "https://arxiv.org/abs/2507.17962",
        "doi_arxiv_id": "arXiv:2507.17962",
        "notes": "Architecture-specific RAG knowledge base makes it more relevant than generic RTL generation.",
    },
    {
        "id": "searchgem5_2023",
        "status": "main",
        "title": "SearchGEM5: Towards Reliable gem5 with Search Based Software Testing and Large Language Models",
        "authors": "Aidan Dakhama; Karine Even-Mendoza; William B. Langdon; Hector Menendez; Justyna Petke",
        "year": "2023",
        "date": "2023-12-08",
        "venue": "SSBSE 2023 / LNCS",
        "paper_type": "conference paper",
        "direction": "LLM-assisted architecture simulator testing",
        "technical_route": "LLM for simulation workflow",
        "llm_role": "ChatGPT parameterizes C programs that seed a custom AFL++ fuzzing flow",
        "architecture_target": "gem5 system simulator reliability",
        "simulator_toolchain": "gem5; AFL++; differential testing",
        "evaluation_metrics": "4,005 binaries; +1000 lines test coverage; 244 divergent gem5 simulation instances",
        "category": "03_LLM_for_Simulation_Modeling_and_gem5",
        "recent_half_year_arxiv": "no",
        "official_url": "https://doi.org/10.1007/978-3-031-48796-5_14",
        "oa_url": "https://solar.cs.ucl.ac.uk/pdf/Dakhama_2023_SSBSE.pdf",
        "doi_arxiv_id": "DOI:10.1007/978-3-031-48796-5_14",
        "notes": "Not an architecture design paper, but directly improves architecture simulator reliability using LLMs.",
    },
    {
        "id": "searchsys_2025",
        "status": "main",
        "title": "Enhancing Search-Based Testing with LLMs for Finding Bugs in System Simulators",
        "authors": "Aidan Dakhama; Karine Even-Mendoza; William B. Langdon; Hector D. Menendez; Justyna Petke",
        "year": "2025",
        "date": "2025-07-10",
        "venue": "Automated Software Engineering",
        "paper_type": "journal article",
        "direction": "LLM-assisted architecture simulator testing",
        "technical_route": "LLM for simulation workflow",
        "llm_role": "Six LLMs generate structured test inputs that seed customized AFL++ mutation and differential testing",
        "architecture_target": "gem5 system simulator reliability",
        "simulator_toolchain": "gem5; AFL++; SearchSYS; differential testing",
        "evaluation_metrics": "21 new gem5 bugs; 14 missimulations; 7 crashes; 101,442 issue-triggering inputs",
        "category": "03_LLM_for_Simulation_Modeling_and_gem5",
        "recent_half_year_arxiv": "no",
        "official_url": "https://doi.org/10.1007/s10515-025-00531-7",
        "oa_url": "https://solar.cs.ucl.ac.uk/pdf/ASE_J_Gem5_11-apr-2025.pdf",
        "doi_arxiv_id": "DOI:10.1007/s10515-025-00531-7",
        "notes": "Expanded journal version of SearchGEM5; strong simulator/tool reliability item.",
    },
    {
        "id": "looprag_2026",
        "status": "main",
        "title": "LOOPRAG: Enhancing Loop Transformation Optimization with Retrieval-Augmented Large Language Models",
        "authors": "Yijie Zhi; Yayu Cao; Jianhua Dai; Xiaoyang Han; Jingwen Pu; Qingran Wu; Sheng Cheng; Ming Cai",
        "year": "2026",
        "date": "2026-03-24",
        "venue": "ASPLOS 2026",
        "paper_type": "top conference paper",
        "direction": "compiler optimization with RAG LLMs",
        "technical_route": "Cross-layer/HPC optimization",
        "llm_role": "RAG augments LLM loop-transformation optimizer",
        "architecture_target": "loop transformation and performance portability; architecture-adjacent compiler optimization",
        "simulator_toolchain": "not available from public program page",
        "evaluation_metrics": "not available from public program page",
        "category": "04_LLM_for_Performance_Analysis_and_Optimization",
        "recent_half_year_arxiv": "yes",
        "official_url": "https://doi.org/10.1145/3779212.3790183",
        "oa_url": "https://arxiv.org/abs/2512.15766",
        "doi_arxiv_id": "arXiv:2512.15766; DOI:10.1145/3779212.3790183",
        "notes": "Top-venue ASPLOS item; compiler/HPC-adjacent rather than architecture design, retained due broad scope.",
    },
    {
        "id": "lpo_2026",
        "status": "main",
        "title": "LPO: Discovering Missed Peephole Optimizations with Large Language Models",
        "authors": "Zhenyang Xu; Hongxu Xu; Yongqiang Tian; Xintong Zhou; Chengnian Sun",
        "year": "2026",
        "date": "2026-03-24",
        "venue": "ASPLOS 2026",
        "paper_type": "top conference paper",
        "direction": "compiler optimization discovery with LLMs",
        "technical_route": "Cross-layer/HPC optimization",
        "llm_role": "LLM discovers missed peephole optimizations",
        "architecture_target": "compiler optimization; systems/compiler layer adjacent to architecture",
        "simulator_toolchain": "not available from public program page",
        "evaluation_metrics": "not available from public program page",
        "category": "04_LLM_for_Performance_Analysis_and_Optimization",
        "recent_half_year_arxiv": "no",
        "official_url": "https://www.asplos-conference.org/asplos2026/program/index.html",
        "oa_url": "https://arxiv.org/abs/2508.16125",
        "doi_arxiv_id": "arXiv:2508.16125",
        "notes": "Included as ASPLOS top-venue LLM-for-systems optimization, but not core computer architecture.",
    },
    {
        "id": "optimas_2026",
        "status": "main",
        "title": "Optimas: An Intelligent Analytics-Informed Generative AI Framework for Performance Optimization",
        "authors": "Mohammad Zaeed; Tanzima Z. Islam; Vladimir Indic",
        "year": "2026",
        "date": "2026-04",
        "venue": "arXiv",
        "paper_type": "recent preprint",
        "direction": "LLM agents for GPU/HPC performance optimization",
        "technical_route": "Cross-layer/HPC optimization",
        "llm_role": "Multi-agent workflow maps performance diagnostics to literature-backed code transformations and validates execution",
        "architecture_target": "NVIDIA GPU code/performance optimization",
        "simulator_toolchain": "profilers/reports; code generation; execution validation",
        "evaluation_metrics": "3,410 experiments; 100% correct code; improves performance in 98.82% experiments; 8.02%-79.09% average gains",
        "category": "06_LLM_for_HPC_System_Codesign",
        "recent_half_year_arxiv": "yes",
        "official_url": "https://arxiv.org/abs/2604.23892",
        "oa_url": "https://arxiv.org/abs/2604.23892",
        "doi_arxiv_id": "arXiv:2604.23892",
        "notes": "Recent LLM-agent HPC/GPU performance optimizer; architecture-adjacent rather than classic Arch venue.",
    },
    {
        "id": "swizzleperf_2025",
        "status": "main",
        "title": "SwizzlePerf: Hardware-aware LLMs for GPU Kernel Performance Optimization",
        "authors": "Arya Tschand; Muhammad Awad; Ryan Swann; Kesavan Ramakrishnan; Jeffrey Ma; Keith Lowery; Ganesh Dasika; Vijay Janapa Reddi",
        "year": "2025",
        "date": "2025-08",
        "venue": "arXiv",
        "paper_type": "preprint",
        "direction": "hardware-aware LLMs for GPU kernel performance",
        "technical_route": "Cross-layer/HPC optimization",
        "llm_role": "LLM assists hardware-aware kernel performance optimization",
        "architecture_target": "GPU kernels and hardware-aware code transformation",
        "simulator_toolchain": "not resolved in this pass",
        "evaluation_metrics": "not resolved in this pass",
        "category": "06_LLM_for_HPC_System_Codesign",
        "recent_half_year_arxiv": "no",
        "official_url": "https://arxiv.org/abs/2508.20258",
        "oa_url": "https://arxiv.org/abs/2508.20258",
        "doi_arxiv_id": "arXiv:2508.20258",
        "notes": "Mentioned in QuArch references; include as architecture-adjacent GPU performance work.",
    },
    {
        "id": "hardware_llm_scoping_2025",
        "status": "main",
        "title": "Hardware Design and Verification with Large Language Models: A Scoping Review, Challenges, and Open Issues",
        "authors": "Meisam Abdollahi; Seyedeh Faegheh Yeganli; Mohammad (Amir) Baharloo; Amirali Baniasadi",
        "year": "2025",
        "date": "2024-12-30",
        "venue": "Electronics",
        "paper_type": "survey",
        "direction": "survey of LLMs for hardware design and verification",
        "technical_route": "Survey/vision",
        "llm_role": "Reviews LLM use across specification, RTL, verification, chip architecture design, and EDA workflows",
        "architecture_target": "chip architecture; RTL; verification; physical design",
        "simulator_toolchain": "survey",
        "evaluation_metrics": "taxonomy and open issues",
        "category": "90_Surveys_Position_and_Vision",
        "recent_half_year_arxiv": "no",
        "official_url": "https://www.mdpi.com/2079-9292/14/1/120",
        "oa_url": "https://www.mdpi.com/2079-9292/14/1/120/pdf",
        "doi_arxiv_id": "DOI:10.3390/electronics14010120",
        "notes": "Broad EDA/hardware survey; useful for taxonomy, not core Arch venue.",
    },
    {
        "id": "ai4eda_workshop_report_2024",
        "status": "main",
        "title": "Report for NSF Workshop on AI for Electronic Design Automation",
        "authors": "Deming Chen; Vijay Ganesh; Weikai Li; Yizhou Sun",
        "year": "2024",
        "date": "2024-12",
        "venue": "NSF Workshop Report",
        "paper_type": "report",
        "direction": "AI/LLM for EDA roadmap",
        "technical_route": "Survey/vision",
        "llm_role": "Discusses LLMs, GNNs, RL and AI methods for HLS/LLS, optimization, physical synthesis, test and verification",
        "architecture_target": "EDA and hardware design flows",
        "simulator_toolchain": "workshop report",
        "evaluation_metrics": "recommendations and research agenda",
        "category": "90_Surveys_Position_and_Vision",
        "recent_half_year_arxiv": "no",
        "official_url": "https://ai4eda-workshop.github.io/",
        "oa_url": "",
        "doi_arxiv_id": "",
        "notes": "Not strictly LLM-only; included as context for EDA/HLS technical route.",
    },
    {
        "id": "chipnemo_2023",
        "status": "borderline",
        "title": "ChipNeMo: Domain-Adapted LLMs for Chip Design",
        "authors": "NVIDIA research team",
        "year": "2023",
        "date": "2023-11",
        "venue": "arXiv",
        "paper_type": "preprint",
        "direction": "domain-adapted LLMs for chip design",
        "technical_route": "Hardware design/spec optimization; RAG/assistant",
        "llm_role": "Domain-adapted LLM assistant for chip-design tasks",
        "architecture_target": "chip design workflows; EDA assistance",
        "simulator_toolchain": "chip-design corpora; domain adaptation; retrieval",
        "evaluation_metrics": "domain task quality metrics; details in paper",
        "category": "99_Candidates_Borderline",
        "recent_half_year_arxiv": "no",
        "official_url": "https://arxiv.org/abs/2311.00176",
        "oa_url": "https://arxiv.org/abs/2311.00176",
        "doi_arxiv_id": "arXiv:2311.00176",
        "notes": "Important chip-design LLM paper, but broader EDA rather than computer architecture.",
    },
    {
        "id": "verigen_2024",
        "status": "borderline",
        "title": "Verigen: A Large Language Model for Verilog Code Generation",
        "authors": "Shailja Thakur; Baleegh Ahmad; Hammond Pearce; Benjamin Tan; Brendan Dolan-Gavitt; Ramesh Karri; Siddharth Garg",
        "year": "2024",
        "date": "2024",
        "venue": "ACM Transactions on Design Automation of Electronic Systems",
        "paper_type": "journal article",
        "direction": "LLM for Verilog generation",
        "technical_route": "Hardware design/spec optimization",
        "llm_role": "Fine-tuned LLM generates Verilog code",
        "architecture_target": "RTL code generation",
        "simulator_toolchain": "Verilog evaluation/testbench flow",
        "evaluation_metrics": "Verilog generation correctness metrics",
        "category": "99_Candidates_Borderline",
        "recent_half_year_arxiv": "no",
        "official_url": "https://doi.org/10.1145/3643681",
        "oa_url": "https://arxiv.org/abs/2308.00708",
        "doi_arxiv_id": "arXiv:2308.00708; DOI:10.1145/3643681",
        "notes": "Pure RTL-generation baseline; included as borderline due EDA extension.",
    },
    {
        "id": "rtllm_2024",
        "status": "borderline",
        "title": "RTLLM: An Open-Source Benchmark for Design RTL Generation with Large Language Model",
        "authors": "Yao Lu; Shang Liu; Qijun Zhang; Zhiyao Xie",
        "year": "2024",
        "date": "2024-01",
        "venue": "ASP-DAC 2024",
        "paper_type": "conference paper",
        "direction": "benchmark for LLM RTL generation",
        "technical_route": "Benchmark/dataset; hardware design/spec optimization",
        "llm_role": "Evaluates LLMs for RTL design generation",
        "architecture_target": "RTL generation",
        "simulator_toolchain": "RTL design benchmark and validation flow",
        "evaluation_metrics": "RTL generation benchmark pass/fail metrics",
        "category": "99_Candidates_Borderline",
        "recent_half_year_arxiv": "no",
        "official_url": "https://doi.org/10.1109/ASP-DAC58780.2024.10473904",
        "oa_url": "",
        "doi_arxiv_id": "DOI:10.1109/ASP-DAC58780.2024.10473904",
        "notes": "Strong EDA benchmark; not architecture-specific but useful as foundation.",
    },
    {
        "id": "rtlcoder_2023",
        "status": "borderline",
        "title": "RTLCoder: Outperforming GPT-3.5 in Design RTL Generation with Our Open-Source Dataset and Lightweight Solution",
        "authors": "Shang Liu; Wenji Fang; Yao Lu; Qijun Zhang; Hongce Zhang; Zhiyao Xie",
        "year": "2023",
        "date": "2023-12",
        "venue": "arXiv",
        "paper_type": "preprint",
        "direction": "LLM for RTL generation",
        "technical_route": "Hardware design/spec optimization",
        "llm_role": "Lightweight model/dataset for RTL design generation",
        "architecture_target": "RTL generation",
        "simulator_toolchain": "RTL generation dataset/evaluation",
        "evaluation_metrics": "outperforms GPT-3.5 on RTL generation tasks",
        "category": "99_Candidates_Borderline",
        "recent_half_year_arxiv": "no",
        "official_url": "https://arxiv.org/abs/2312.08617",
        "oa_url": "https://arxiv.org/abs/2312.08617",
        "doi_arxiv_id": "arXiv:2312.08617",
        "notes": "Pure RTL-generation; included as candidate.",
    },
    {
        "id": "spec_llm_2024",
        "status": "borderline",
        "title": "SpecLLM: Exploring Generation and Review of VLSI Design Specification with Large Language Model",
        "authors": "Mingjie Li; Wenji Fang; Qijun Zhang; Zhiyao Xie",
        "year": "2024",
        "date": "2024-01",
        "venue": "arXiv",
        "paper_type": "preprint",
        "direction": "LLM for VLSI specification generation/review",
        "technical_route": "Hardware design/spec optimization",
        "llm_role": "LLM generates and reviews VLSI design specifications",
        "architecture_target": "VLSI/SoC design specification",
        "simulator_toolchain": "specification evaluation flow",
        "evaluation_metrics": "spec generation/review quality",
        "category": "99_Candidates_Borderline",
        "recent_half_year_arxiv": "no",
        "official_url": "https://arxiv.org/abs/2401.13266",
        "oa_url": "https://arxiv.org/abs/2401.13266",
        "doi_arxiv_id": "arXiv:2401.13266",
        "notes": "Specification-level rather than architecture evaluation; include as borderline.",
    },
    {
        "id": "from_english_to_asic_2024",
        "status": "borderline",
        "title": "From English to ASIC: Hardware Implementation with Large Language Model",
        "authors": "E. Goh; M. Xiang; I. Wey; T. H. Teo",
        "year": "2024",
        "date": "2024-03",
        "venue": "arXiv",
        "paper_type": "preprint",
        "direction": "natural-language to ASIC/RTL design",
        "technical_route": "Hardware design/spec optimization",
        "llm_role": "LLM maps English design intent toward hardware implementation",
        "architecture_target": "ASIC/RTL implementation",
        "simulator_toolchain": "RTL/ASIC tool flow",
        "evaluation_metrics": "implementation correctness/quality",
        "category": "99_Candidates_Borderline",
        "recent_half_year_arxiv": "no",
        "official_url": "https://arxiv.org/abs/2403.07039",
        "oa_url": "https://arxiv.org/abs/2403.07039",
        "doi_arxiv_id": "arXiv:2403.07039",
        "notes": "Natural-language hardware implementation; kept as EDA candidate.",
    },
    {
        "id": "advanced_verilog_ppa_2023",
        "status": "borderline",
        "title": "Advanced Language Model-Driven Verilog Development: Enhancing Power, Performance, and Area Optimization in Code Synthesis",
        "authors": "Keshav Thorat; Jason Zhao; Yucheng Liu; Hengyue Peng; Xuechao Xie; Bolei Lei; Zhiru Zhang; Chao Ding",
        "year": "2023",
        "date": "2023-12",
        "venue": "arXiv",
        "paper_type": "preprint",
        "direction": "LLM-driven Verilog PPA optimization",
        "technical_route": "Hardware design/spec optimization",
        "llm_role": "LLM assists Verilog code synthesis with PPA objectives",
        "architecture_target": "RTL/PPA",
        "simulator_toolchain": "Verilog synthesis/evaluation flow",
        "evaluation_metrics": "PPA optimization metrics",
        "category": "99_Candidates_Borderline",
        "recent_half_year_arxiv": "no",
        "official_url": "https://arxiv.org/abs/2312.01022",
        "oa_url": "https://arxiv.org/abs/2312.01022",
        "doi_arxiv_id": "arXiv:2312.01022",
        "notes": "PPA angle makes it somewhat architecture-relevant, but still mostly RTL/EDA.",
    },
]


ARXIV_NS = "{http://www.w3.org/2005/Atom}"


def safe_filename(text: str, max_len: int = 64) -> str:
    text = re.sub(r"[^A-Za-z0-9]+", "_", text).strip("_")
    return text[:max_len].strip("_") or "paper"


def arxiv_id(entry):
    m = re.search(r"arXiv:(\d{4}\.\d{4,5})", entry.get("doi_arxiv_id", ""))
    return m.group(1) if m else ""


def first_author(entry):
    authors = entry.get("authors", "")
    if not authors or authors.startswith("authors not"):
        return "Unknown"
    first = authors.split(";")[0].strip()
    last = first.split()[-1] if first else "Unknown"
    return safe_filename(last, 24)


def output_pdf_name(entry):
    venue = safe_filename(entry.get("venue", "unknown").split("/")[0].strip(), 20)
    title = safe_filename(entry["title"], 48)
    return f"{entry['year']}_{venue}_{first_author(entry)}_{title}.pdf"


def ensure_dirs():
    ROOT.mkdir(exist_ok=True)
    for cat in CATEGORIES:
        (ROOT / cat).mkdir(exist_ok=True)
    (ROOT / "metadata").mkdir(exist_ok=True)
    (ROOT / "scripts").mkdir(exist_ok=True)


def request_url(url, timeout=60):
    req = urllib.request.Request(
        url,
        headers={
            "User-Agent": "Codex literature collection for academic metadata/PDF download (mailto:research@example.com)"
        },
    )
    return urllib.request.urlopen(req, timeout=timeout)


def enrich_arxiv(entries):
    if os.environ.get("FETCH_ARXIV_METADATA") != "1":
        return {}
    ids = sorted({arxiv_id(e) for e in entries if arxiv_id(e)})
    metadata = {}
    for i in range(0, len(ids), 20):
        batch = ids[i : i + 20]
        url = "https://export.arxiv.org/api/query?id_list=" + ",".join(batch)
        try:
            with request_url(url, timeout=60) as resp:
                data = resp.read()
            root = ET.fromstring(data)
            for item in root.findall(f"{ARXIV_NS}entry"):
                paper_id = item.findtext(f"{ARXIV_NS}id", default="").rsplit("/", 1)[-1]
                paper_id = paper_id.split("v")[0]
                metadata[paper_id] = {
                    "arxiv_title": " ".join(item.findtext(f"{ARXIV_NS}title", default="").split()),
                    "arxiv_authors": "; ".join(
                        " ".join(a.findtext(f"{ARXIV_NS}name", default="").split())
                        for a in item.findall(f"{ARXIV_NS}author")
                    ),
                    "published": item.findtext(f"{ARXIV_NS}published", default="")[:10],
                    "updated": item.findtext(f"{ARXIV_NS}updated", default="")[:10],
                    "summary": " ".join(item.findtext(f"{ARXIV_NS}summary", default="").split()),
                }
        except Exception as exc:
            print(f"arXiv metadata failed for {batch}: {exc}", file=sys.stderr)
        time.sleep(3)
    for e in entries:
        aid = arxiv_id(e)
        if aid in metadata:
            m = metadata[aid]
            if not e.get("date") or len(e["date"]) < 8:
                e["date"] = m["published"]
            if e.get("authors", "").startswith("authors not"):
                e["authors"] = m["arxiv_authors"]
            e["notes"] = (e.get("notes", "") + f" arXiv published {m['published']}; updated {m['updated']}.").strip()
    return metadata


def download_file(url, dest):
    try:
        with request_url(url, timeout=90) as resp:
            ctype = resp.headers.get("Content-Type", "")
            data = resp.read()
        if len(data) < 1024:
            return False, f"too small ({len(data)} bytes)"
        dest.write_bytes(data)
        return True, f"downloaded {len(data)} bytes; content-type={ctype}"
    except urllib.error.HTTPError as exc:
        return False, f"HTTP {exc.code}: {exc.reason}"
    except Exception as exc:
        return False, str(exc)


def existing_pdf_is_valid(path):
    if not path.exists() or path.stat().st_size < 1024:
        return False
    with path.open("rb") as f:
        return f.read(5) == b"%PDF-"


def clean_stale_pdfs(entries):
    expected = {
        (ROOT / e["category"] / output_pdf_name(e)).resolve()
        for e in entries
    }
    for category in CATEGORIES:
        category_dir = ROOT / category
        for pdf in category_dir.glob("*.pdf"):
            if pdf.resolve() not in expected:
                pdf.unlink()


def download_pdfs(entries):
    log_rows = []
    for e in entries:
        dest = ROOT / e["category"] / output_pdf_name(e)
        url = ""
        aid = arxiv_id(e)
        if aid:
            url = f"https://arxiv.org/pdf/{aid}"
        elif e.get("oa_url", "").lower().endswith(".pdf"):
            url = e["oa_url"]
        elif e.get("oa_url", "").endswith("/pdf"):
            url = e["oa_url"]
        if url:
            if existing_pdf_is_valid(dest):
                ok, msg = True, f"reused existing PDF ({dest.stat().st_size} bytes)"
            else:
                ok, msg = download_file(url, dest)
            if ok:
                e["pdf_path"] = str(dest)
                log_rows.append({**{k: e.get(k, "") for k in FIELDS}, "download_status": "success", "download_url": url, "download_message": msg})
            else:
                e["pdf_path"] = ""
                log_rows.append({**{k: e.get(k, "") for k in FIELDS}, "download_status": "failed", "download_url": url, "download_message": msg})
        else:
            e["pdf_path"] = ""
            log_rows.append({**{k: e.get(k, "") for k in FIELDS}, "download_status": "not_attempted", "download_url": "", "download_message": "no open PDF URL found in this pass"})
        if url and not existing_pdf_is_valid(dest):
            time.sleep(1)
    return log_rows


def write_csv(path, rows, fields):
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def bib_key(entry):
    return f"{first_author(entry).lower()}{entry.get('year','')}{safe_filename(entry['title'].split(':')[0], 20).lower()}"


def write_bib(entries):
    lines = []
    for e in entries:
        entry_type = "article"
        venue = e.get("venue", "")
        if any(v in venue for v in ["ASPLOS", "DAC", "SSBSE", "COINS", "ASP-DAC", "CAMS"]):
            entry_type = "inproceedings"
        if "arXiv" in venue and "/" not in venue:
            entry_type = "misc"
        lines.append(f"@{entry_type}{{{bib_key(e)},")
        lines.append(f"  title = {{{e['title']}}},")
        lines.append(f"  author = {{{e.get('authors','').replace(';', ' and')}}},")
        lines.append(f"  year = {{{e.get('year','')}}},")
        if e.get("venue"):
            lines.append(f"  note = {{{e.get('venue','')}}},")
        if "arXiv:" in e.get("doi_arxiv_id", ""):
            lines.append(f"  eprint = {{{arxiv_id(e)}}},")
            lines.append("  archivePrefix = {arXiv},")
        if "DOI:" in e.get("doi_arxiv_id", ""):
            doi = e["doi_arxiv_id"].split("DOI:", 1)[1].split(";")[0].strip()
            lines.append(f"  doi = {{{doi}}},")
        url = e.get("official_url") or e.get("oa_url")
        if url:
            lines.append(f"  url = {{{url}}},")
        lines.append("}\n")
    (ROOT / "references.bib").write_text("\n".join(lines), encoding="utf-8")


def write_review(entries):
    main = [e for e in entries if e["status"] == "main"]
    borderline = [e for e in entries if e["status"] != "main"]
    by_cat = {}
    for e in main:
        by_cat.setdefault(e["category"], []).append(e)

    lines = [
        "# LLM for Computer Architecture 文献分类综述",
        "",
        f"生成日期：2026-04-30。本库从 2022-11-30 之后开始收录，只纳入实质使用 LLM/基础模型/多模态大模型/Agent/RAG 的工作。当前主集 {len(main)} 篇，边界候选 {len(borderline)} 篇。",
        "",
        "## 1. 总体判断",
        "",
        "这个方向仍然很新。严格意义上的顶级体系结构 venue 中，已能看到 ASPLOS 2026 的 PF-LLM、CacheMind、LOOPRAG、LPO 等条目，但 ISCA/MICRO/HPCA 近年更多是 Arch for LLM，而不是 LLM for Arch。当前 LLM for Arch 的主阵地集中在 arXiv、HLS/EDA、workshop，以及少量 ASPLOS/期刊论文。",
        "",
        "## 2. 技术路线分类",
        "",
    ]
    for cat, desc in CATEGORIES.items():
        if cat == "99_Candidates_Borderline":
            continue
        papers = by_cat.get(cat, [])
        if not papers:
            continue
        lines.append(f"### {cat}: {desc}")
        for e in papers:
            path = e.get("pdf_path", "")
            path_note = f"PDF: `{path}`" if path else "PDF: 未下载或未找到开放版本"
            lines.append(
                f"- **{e['title']}** ({e['year']}, {e['venue']}): {e['technical_route']}。"
                f"目标：{e['architecture_target']}；LLM 角色：{e['llm_role']}。{path_note}"
            )
        lines.append("")

    lines += [
        "## 3. 代表性趋势",
        "",
        "- **从聊天助手到闭环代理**：gem5 Co-Pilot、LLM-DSE、iDSE、Agentic Architect 都把 LLM 放进“生成候选设计-调用工具-读取反馈-再生成”的闭环。",
        "- **从性能数字到可解释原因**：CacheMind 这类 RAG/trace-grounded 方法开始让 LLM 面向 cache trace、PC、地址和策略行为做可验证解释。",
        "- **HLS 是最活跃落点**：因为 HLS 的 pragma/DSE 空间巨大且评估昂贵，LLM 的知识推理、代码理解和探索策略比较容易形成闭环。",
        "- **严格微架构策略设计刚起步**：PF-LLM、Agentic Architect、CacheMind 是最贴近 cache/prefetch/branch predictor 的方向，数量还少但增长快。",
        "- **评测基础正在出现**：QuArch 开始系统测 LLM 的体系结构知识、分析、设计和实现能力。",
        "",
        "## 4. 最近半年 arXiv 重点",
        "",
    ]
    recent = [e for e in main if e.get("recent_half_year_arxiv") == "yes"]
    for e in recent:
        lines.append(f"- {e['date']}：**{e['title']}**，{e['direction']}。")
    lines += [
        "",
        "## 5. 边界候选说明",
        "",
        "以下论文是 EDA/RTL/Verilog/规范生成方向，未必属于 computer architecture 主集，但对构建 LLM-for-chip/architecture 工具链很有参考价值：",
        "",
    ]
    for e in borderline:
        lines.append(f"- **{e['title']}** ({e['year']}): {e['notes']}")
    lines += [
        "",
        "## 6. 明确排除原则",
        "",
        "- 只讨论如何加速/部署 LLM 的硬件架构论文，如果没有用 LLM 来辅助设计或优化，默认排除。",
        "- 传统 ML/RL/BO/GNN-only 的 AI for Architecture 工作不进入本库主集。",
        "- 纯 RTL 代码生成论文放在边界候选，除非显式涉及架构级 DSE、PPA tradeoff、HLS pragma 或微架构策略。",
    ]
    (ROOT / "taxonomy_review.md").write_text("\n".join(lines), encoding="utf-8")


def write_xlsx(entries, log_rows):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception as exc:
        print(f"openpyxl unavailable, skipping xlsx: {exc}", file=sys.stderr)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "literature_index"
    ws.append(FIELDS)
    for e in entries:
        ws.append([e.get(f, "") for f in FIELDS])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    for idx, field in enumerate(FIELDS, start=1):
        width = 18
        if field in {"title", "authors", "notes", "llm_role", "evaluation_metrics"}:
            width = 45
        elif field in {"official_url", "oa_url", "pdf_path"}:
            width = 38
        ws.column_dimensions[get_column_letter(idx)].width = width

    log_ws = wb.create_sheet("download_log")
    log_fields = FIELDS + ["download_status", "download_url", "download_message"]
    log_ws.append(log_fields)
    for row in log_rows:
        log_ws.append([row.get(f, "") for f in log_fields])
    for cell in log_ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FCE4D6")
    log_ws.auto_filter.ref = log_ws.dimensions
    log_ws.freeze_panes = "A2"
    for idx, field in enumerate(log_fields, start=1):
        width = 18
        if field in {"title", "authors", "notes", "download_message"}:
            width = 45
        elif field in {"official_url", "oa_url", "pdf_path", "download_url"}:
            width = 38
        log_ws.column_dimensions[get_column_letter(idx)].width = width

    wb.save(ROOT / "literature_index.xlsx")


def md_escape(value):
    text = str(value or "").strip()
    if not text:
        return "-"
    return text.replace("\\", "\\\\").replace("|", "\\|").replace("\r\n", "<br>").replace("\n", "<br>")


def md_link(label, target):
    if not target:
        return "-"
    return f"[{md_escape(label)}](<{target}>)"


def pdf_markdown_link(entry):
    path = entry.get("pdf_path", "")
    if not path:
        return "-"
    pdf_path = Path(path)
    try:
        rel_path = pdf_path.relative_to(ROOT)
    except ValueError:
        parts = pdf_path.parts
        if parts and parts[0] == ROOT.name:
            rel_path = Path(*parts[1:])
        else:
            rel_path = pdf_path
    return md_link("PDF", rel_path.as_posix())


def official_oa_markdown_links(entry):
    links = []
    official = entry.get("official_url", "")
    oa = entry.get("oa_url", "")
    if official:
        links.append(md_link("Official", official))
    if oa and oa != official:
        links.append(md_link("OA", oa))
    return " / ".join(links) if links else "-"


def status_label(entry):
    labels = [entry.get("status", "")]
    if entry.get("recent_half_year_arxiv") == "yes":
        labels.append("recent arXiv")
    return "; ".join(label for label in labels if label)


def sorted_entries(entries):
    category_order = {category: idx for idx, category in enumerate(CATEGORIES)}
    return sorted(
        entries,
        key=lambda e: (
            category_order.get(e.get("category", ""), 99),
            -int(e.get("year") or 0),
            e.get("title", "").lower(),
        ),
    )


def write_literature_index_by_category(entries):
    lines = [
        "# Literature Index by Category",
        "",
        "Compact Markdown view for quick browsing. PDF links are local relative paths from this directory.",
        "",
        f"- Total records: {len(entries)}",
        f"- Main records: {sum(1 for e in entries if e.get('status') == 'main')}",
        f"- Borderline records: {sum(1 for e in entries if e.get('status') != 'main')}",
        "",
    ]
    by_category = {}
    for entry in sorted_entries(entries):
        by_category.setdefault(entry.get("category", ""), []).append(entry)

    for category, description in CATEGORIES.items():
        papers = by_category.get(category, [])
        if not papers:
            continue
        lines += [
            f"## {category}",
            "",
            f"{description}. Records: {len(papers)}.",
            "",
            "| Title | Year | Venue | Technical Route | Target | PDF | Official/OA | Status |",
            "|---|---:|---|---|---|---|---|---|",
        ]
        for entry in papers:
            lines.append(
                "| "
                + " | ".join(
                    [
                        md_escape(entry.get("title")),
                        md_escape(entry.get("year")),
                        md_escape(entry.get("venue")),
                        md_escape(entry.get("technical_route")),
                        md_escape(entry.get("architecture_target")),
                        pdf_markdown_link(entry),
                        official_oa_markdown_links(entry),
                        md_escape(status_label(entry)),
                    ]
                )
                + " |"
            )
        lines.append("")

    (ROOT / "literature_index_by_category.md").write_text("\n".join(lines), encoding="utf-8")


def write_literature_cards(entries):
    lines = [
        "# Literature Cards",
        "",
        "Detailed per-paper Markdown view. Use `literature_index_by_category.md` for quick scanning.",
        "",
    ]
    by_category = {}
    for entry in sorted_entries(entries):
        by_category.setdefault(entry.get("category", ""), []).append(entry)

    card_number = 1
    for category, description in CATEGORIES.items():
        papers = by_category.get(category, [])
        if not papers:
            continue
        lines += [f"## {category}", "", description + ".", ""]
        for entry in papers:
            lines += [
                f"### {card_number}. {md_escape(entry.get('title'))}",
                "",
                f"- **ID:** `{md_escape(entry.get('id'))}`",
                f"- **Authors:** {md_escape(entry.get('authors'))}",
                f"- **Year / Venue:** {md_escape(entry.get('year'))} / {md_escape(entry.get('venue'))}",
                f"- **Paper type:** {md_escape(entry.get('paper_type'))}",
                f"- **Status:** {md_escape(status_label(entry))}",
                f"- **Direction:** {md_escape(entry.get('direction'))}",
                f"- **Technical route:** {md_escape(entry.get('technical_route'))}",
                f"- **LLM role:** {md_escape(entry.get('llm_role'))}",
                f"- **Architecture target:** {md_escape(entry.get('architecture_target'))}",
                f"- **Simulator/toolchain:** {md_escape(entry.get('simulator_toolchain'))}",
                f"- **Evaluation metrics:** {md_escape(entry.get('evaluation_metrics'))}",
                f"- **PDF:** {pdf_markdown_link(entry)}",
                f"- **Official/OA:** {official_oa_markdown_links(entry)}",
                f"- **DOI/arXiv:** {md_escape(entry.get('doi_arxiv_id'))}",
                f"- **Notes:** {md_escape(entry.get('notes'))}",
                "",
            ]
            card_number += 1

    (ROOT / "literature_cards.md").write_text("\n".join(lines), encoding="utf-8")


def write_download_summary(log_rows):
    status_counts = {}
    for row in log_rows:
        status_counts[row.get("download_status", "")] = status_counts.get(row.get("download_status", ""), 0) + 1

    lines = [
        "# Download Summary",
        "",
        "PDF download status generated from `download_log.csv`.",
        "",
        "| Download status | Count |",
        "|---|---:|",
    ]
    for status in ["success", "failed", "not_attempted"]:
        lines.append(f"| {md_escape(status)} | {status_counts.get(status, 0)} |")

    missing = [row for row in log_rows if row.get("download_status") != "success"]
    lines += [
        "",
        "## Not Downloaded",
        "",
        "| ID | Title | Status | Reason | Official/OA |",
        "|---|---|---|---|---|",
    ]
    for row in missing:
        lines.append(
            "| "
            + " | ".join(
                [
                    f"`{md_escape(row.get('id'))}`",
                    md_escape(row.get("title")),
                    md_escape(row.get("download_status")),
                    md_escape(row.get("download_message")),
                    official_oa_markdown_links(row),
                ]
            )
            + " |"
        )

    lines += [
        "",
        "## Downloaded PDFs",
        "",
        "| ID | Title | Local PDF |",
        "|---|---|---|",
    ]
    for row in log_rows:
        if row.get("download_status") == "success":
            lines.append(
                "| "
                + " | ".join(
                    [
                        f"`{md_escape(row.get('id'))}`",
                        md_escape(row.get("title")),
                        pdf_markdown_link(row),
                    ]
                )
                + " |"
            )

    (ROOT / "download_summary.md").write_text("\n".join(lines), encoding="utf-8")


def write_markdown_views(entries, log_rows):
    write_literature_index_by_category(entries)
    write_literature_cards(entries)
    write_download_summary(log_rows)


def write_readme(entries, log_rows):
    success = sum(1 for r in log_rows if r["download_status"] == "success")
    failed = sum(1 for r in log_rows if r["download_status"] == "failed")
    missing = sum(1 for r in log_rows if r["download_status"] == "not_attempted")
    text = f"""# AI for Arch LLM Literature Library

Scope: LLM/Foundation-Model/Agent/RAG methods for computer architecture, HLS, simulator workflows, and architecture-adjacent chip/system design after 2022-11-30.

Generated: 2026-05-01

Search cutoff: 2026-04-30

- Main records: {sum(1 for e in entries if e['status'] == 'main')}
- Borderline records: {sum(1 for e in entries if e['status'] != 'main')}
- PDF download successes: {success}
- PDF download failures: {failed}
- No open PDF found in this pass: {missing}

Key files:
- `literature_index_by_category.md` - quick grouped Markdown index
- `literature_cards.md` - detailed per-paper Markdown cards
- `download_summary.md` - Markdown download status summary
- `literature_index.csv`
- `literature_index.xlsx`
- `references.bib`
- `taxonomy_review.md`
- `download_log.csv`

Notes:
- Pure "architecture for LLM inference/training" papers were excluded unless the paper uses LLMs to help design, search, verify, or optimize architecture/hardware/software-hardware workflows.
- Closed IEEE/ACM PDFs are not bypassed. When an open version is unavailable, the official DOI/program URL is recorded.
"""
    (ROOT / "README.md").write_text(text, encoding="utf-8")


def main():
    ensure_dirs()
    # Work on a shallow copy to avoid accidental mutation surprises during import.
    entries = [dict(e) for e in ENTRIES]
    enrich_arxiv(entries)
    clean_stale_pdfs(entries)
    log_rows = download_pdfs(entries)
    write_csv(ROOT / "literature_index.csv", entries, FIELDS)
    write_csv(ROOT / "download_log.csv", log_rows, FIELDS + ["download_status", "download_url", "download_message"])
    write_bib(entries)
    write_review(entries)
    write_xlsx(entries, log_rows)
    write_markdown_views(entries, log_rows)
    write_readme(entries, log_rows)
    print(f"Wrote {len(entries)} records to {ROOT}")
    print(f"PDF successes: {sum(1 for r in log_rows if r['download_status'] == 'success')}")


if __name__ == "__main__":
    main()
